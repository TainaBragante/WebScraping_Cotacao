import openpyxl
import time
import random
import os
import shutil
import tempfile
import re
from decimal import Decimal, InvalidOperation
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from openpyxl.styles import NamedStyle, Alignment
from datetime import datetime


# Configurações iniciais
user_data_dir = 'C:\\Users\\Admin\\AppData\\Local\\Google\\Chrome\\User Data'  # Diretório de dados do usuário do Chrome
profile_dir = 'Profile 1'  # Perfil do Chrome "Robô Cotação"
excel_path = 'C:\\Users\\Admin\\Cotacao\\MOEDAS\\MOEDAS.xlsx'  # Caminho do arquivo Excel com as moedas 
dias_para_buscar = [1]  # Dias para buscar os valores de fechamento
temp_dir = None  # Inicializa temp_dir como None
driver = None    # Inicializa driver como None

# Função para pressionar a tecla ESC
def press_esc(driver, times=2):
    actions = ActionChains(driver)
    for _ in range(times):
        actions.send_keys(Keys.ESCAPE).perform()
        time.sleep(1)

# Função para rolar a página para baixo e para cima
def random_scroll(driver):
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(random.uniform(1, 1.5))
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(random.uniform(1, 1.5))

# Função para limpar e converter o valor de fechamento
def parse_close_value(texto: str):
    # Converte strings de preço como '<$0.000000000001' em float 0.000000000001. Retorna None se não houver número válido.
    if not texto:
        return None
    # valores ausentes comuns
    if texto.strip() in {'—', '-', 'N/A'}:
        return None
    # Remove comparadores e caracteres 
    texto_limpo = re.sub(r'^[<>=≤≥≈]*', '', texto)  # Remove comparadores do início
    texto_limpo = re.sub(r'[^\d.,-]', '', texto_limpo) # Remove caracteres 
    texto_limpo = texto_limpo.replace(',', '')  # Remove vírgulas
    # Tenta converter para float
    try:
        return float(Decimal(texto_limpo))
    except (InvalidOperation, ValueError):
        return None

# Configuração do driver do Selenium com perfil do Chrome "Robô Cotação"
def setup_driver_with_profile(user_data_dir, profile_dir):
    temp_dir = tempfile.mkdtemp()
    temp_user_data_dir = os.path.join(temp_dir, 'User Data')
    temp_profile_path = os.path.join(temp_user_data_dir, profile_dir)
    original_profile_path = os.path.join(user_data_dir, profile_dir)
    shutil.copytree(original_profile_path, temp_profile_path)

    chrome_options = Options()
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_argument(f"--user-data-dir={temp_user_data_dir}")
    chrome_options.add_argument(f"--profile-directory={profile_dir}")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.page_load_strategy = 'eager'

    s = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=s, options=chrome_options)
    return driver, temp_dir

# Função para obter os valores de fechamento da tabela
def get_close_values(driver, dias_para_buscar):
    close_values = {}
    try:
        # Aguarda até que a tabela esteja presente
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'table tbody')))
        rows = driver.find_elements(By.CSS_SELECTOR, 'table tbody tr')
        for n_dias_atras in dias_para_buscar:
            index = n_dias_atras - 1  # índice baseado em zero
            if index < len(rows):
                row = rows[index]
                date_cell = row.find_element(By.CSS_SELECTOR, 'td:nth-child(1)').text.strip()
                close_cell = row.find_element(By.CSS_SELECTOR, 'td:nth-child(5)').text.strip()
                close_values[n_dias_atras] = {'date': date_cell, 'close': close_cell}
    except (TimeoutException, NoSuchElementException) as e:
        print(f"Erro ao obter valores de fechamento: {e}")
    return close_values


# Processo principal
try:
    driver, temp_dir = setup_driver_with_profile(user_data_dir, profile_dir)
    book = openpyxl.load_workbook(excel_path)
    coin_page = book['Sheet1']
    
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    value_style = NamedStyle(name='value_style', number_format='0.0000', alignment=Alignment(horizontal='center'))
    
    for row in range(2, coin_page.max_row + 1):
        sheetName = coin_page[f'A{row}'].value
        httpsName = coin_page[f'B{row}'].value
    
        excel_path_save = f'C:\\Users\\Admin\\Cotacao\\COTACAO_DOLAR_COINMARKET\\{sheetName}.xlsx'
        book_save = openpyxl.load_workbook(excel_path_save)
        coin_page_save = book_save['Sheet1']
    
        if 'date_style' not in book_save.style_names:
            book_save.add_named_style(date_style)
        if 'value_style' not in book_save.style_names:
            book_save.add_named_style(value_style)
    
        # Ler datas existentes em um conjunto para verificação mais rápida
        existing_dates = set()
        for cell in coin_page_save['A']:
            if cell.value and isinstance(cell.value, datetime):
                existing_dates.add(cell.value.date())
    
        driver.get(httpsName)
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.TAG_NAME, "body")))
        random_scroll(driver)
        press_esc(driver)
    
        close_values = get_close_values(driver, dias_para_buscar)
    
        for n_dias_atras, data in close_values.items():
            # Converter string de data para datetime.date
            date_str = data['date']
            try:
                data_cotacao = datetime.strptime(date_str, '%b %d, %Y').date()  # Ajuste o formato conforme o site
            except ValueError:
                print(f"Erro ao analisar a data: {date_str}")
                continue
    
            if data_cotacao in existing_dates:
                continue
    
            close_value = data['close']
            close_value_clean = parse_close_value(close_value)
            if close_value_clean is None:
                print(f"Valor de fechamento inválido para {sheetName} na data {data_cotacao}: '{close_value}'")
                continue
            last_row = coin_page_save.max_row + 1
            cell_date = coin_page_save.cell(row=last_row, column=1, value=data_cotacao)
            cell_date.style = 'date_style'
            cell_value = coin_page_save.cell(row=last_row, column=2, value=close_value_clean)
            cell_value.style = 'value_style'
    
        book_save.save(excel_path_save)
        time.sleep(random.uniform(1, 1.5))
finally:
    # Fechar o driver se ele foi iniciado
    if driver is not None:
        driver.quit()
    # Remover o diretório temporário se ele foi criado
    if temp_dir is not None and os.path.exists(temp_dir):
        try:
            shutil.rmtree(temp_dir)
        except Exception as e:
            print(f"Erro ao remover o diretório temporário: {e}")

    time.sleep(random.uniform(4, 8))
